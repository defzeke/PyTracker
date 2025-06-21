from flask import Flask, render_template, request, redirect, url_for, session, make_response, jsonify, send_file, Blueprint
from datetime import datetime, timedelta
from flask_mail import Mail, Message
import random
from dotenv import load_dotenv
import os
import mysql.connector
import calendar
import requests
from abc import ABC, abstractmethod
from werkzeug.utils import secure_filename
from time import time
import pdfkit
import pytz



load_dotenv()
print("API KEY:", os.getenv("OPENWEATHERMAP_API_KEY"))
WKHTMLTOPDF_PATH = r"C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe"  
PDFKIT_CONFIG = pdfkit.configuration(wkhtmltopdf=WKHTMLTOPDF_PATH)
##########################################################################################################################

class Main():
    app = None
    mail = None
    
    def __init__(self):    
        self.app = Flask(__name__)
        Main.app = self.app

        self.app.config['MAIL_SERVER'] = 'smtp.gmail.com'
        self.app.config['MAIL_PORT'] = 465
        self.app.config['MAIL_USERNAME'] = 'pytracker43@gmail.com'
        self.app.config['MAIL_PASSWORD'] = 'ayxg eoaa xkij kiha'
        self.app.config['MAIL_USE_TLS'] = False
        self.app.config['MAIL_USE_SSL'] = True
        self.app.secret_key = '123123123'

        self.app.register_blueprint(OTP.OTP_bp)
        self.app.register_blueprint(Login.login_bp)
        self.app.register_blueprint(User.user_bp)
        self.app.register_blueprint(Funcs.funcs_bp)

        self.mail = Mail(self.app)
        Main.mail = self.mail
    
    def run(self):
        self.app.run(debug=True)

class OTP():
    OTP_bp = Blueprint("OTP", __name__)

    @staticmethod
    def send_otp_email(recipient, otp):
        msg = Message(
            'Your OTP Code',
            sender=("Pytracker", Main.app.config['MAIL_USERNAME']),
            recipients=[recipient]
        )
        msg.body = f'Your OTP is: {otp}'  # Fallback for non-HTML clients
        msg.html = f"""
        <div style="font-family: Arial, sans-serif; max-width: 480px; margin: auto; border: 1px solid #eee; border-radius: 8px; padding: 32px 24px;">
            <h2 style="color: #d32f2f; text-align: center;">Register to Pytracker</h2>
            <p style="font-size: 16px; text-align: center;">Welcome! Enter this code within the next 10 minutes to log in:</p>
            <div style="font-size: 32px; font-weight: bold; letter-spacing: 4px; background: #f5f5f5; border-radius: 6px; padding: 16px; text-align: center; margin: 24px 0;">
                {otp}
            </div>
            <p style="font-size: 12px; color: #888; text-align: center;">
                You’re receiving this email because you requested a register code for your Pytracker account.<br>
                If you did not request this, you can safely ignore this email.
            </p>
            <div style="text-align: center; margin-top: 32px; font-size: 12px; color: #bbb;">
                Made for you with ❤️<br>
                Pytracker Team
            </div>
        </div>
        """
        Main.mail.send(msg)
        
    @OTP_bp.route('/resend_otp', methods=['POST'])
    def resend_otp():
        """
        Handles AJAX request to resend OTP.
        """
        email = session.get('email')
        if not email:
            return jsonify({'success': False, 'message': 'No email found in session.'}), 400
        otp_code = OTP.generate_otp()
        session['otp'] = otp_code
        OTP.send_otp_email(email, otp_code)
        return jsonify({'success': True, 'message': 'A new code has been sent to your email.'})

    @staticmethod
    def generate_otp():
        """Generate a 6-digit numeric OTP as a string."""
        return ''.join([str(random.randint(0, 9)) for _ in range(6)])

#########################################################################################################################


#########################################################################################################################
# CONNECTION FUNCTION
class DB():
    @staticmethod
    def get_mysql_connection():
        return mysql.connector.connect(
            host=os.getenv("MYSQL_HOST"),
            user=os.getenv("MYSQL_USER"),
            password=os.getenv("MYSQL_PASSWORD"),
            database=os.getenv("MYSQL_DATABASE"),
            port=int(os.getenv("MYSQL_PORT"))
        )


#########################################################################################################################
#General functions in User abstract class
class User(ABC):
    user_bp = Blueprint("User", __name__)

    @abstractmethod
    @user_bp.route("/logout:)")
    def logout():
        session.clear()
        return redirect(url_for('Login.login'))

    @abstractmethod
    def get_UI(self):
        return redirect(url_for('Login.login'))
    
    @abstractmethod
    @user_bp.route("/weather")
    def get_weather():
        print("Weather route called!")  
        API_KEY = os.getenv("OPENWEATHERMAP_API_KEY")  
        city = "Manila,PH"
        url = f"https://api.openweathermap.org/data/2.5/weather?q={city}&appid={API_KEY}&units=metric"
        try:
            print("Requesting:", url)  
            resp = requests.get(url, timeout=5)
            data = resp.json()
            print("Weather API response:", data)  
            if data.get("cod") != 200:
                return jsonify({"error": "Weather unavailable"}), 500

            weather = {
                "description": data["weather"][0]["description"].title(),
                "icon": data["weather"][0]["icon"],
                "temp": round(data["main"]["temp"], 1),
                "temp_min": round(data["main"]["temp_min"], 1),
                "temp_max": round(data["main"]["temp_max"], 1),
                "city": data["name"]
            }
            return jsonify(weather)
        except Exception as e:
            print("Exception in /weather:", e)  
            return jsonify({"error": "Weather unavailable"}), 500
        
    
    @abstractmethod
    @user_bp.route("/get_classes")
    def get_classes():
        if 'user' not in session:
            print("No user in session for /get_classes")
            return jsonify([])
        conn = DB.get_mysql_connection()
        cursor = conn.cursor(dictionary=True)
        # Get professor's id_number
        cursor.execute("""
            SELECT id_number FROM login_main
            WHERE TRIM(LOWER(name)) = TRIM(LOWER(%s))
        """, (session['user'],))
        prof = cursor.fetchone()
        if not prof:
            print(f"Professor not found for user: {session['user']}")
            cursor.close()
            conn.close()
            return jsonify([])
        prof_id = prof['id_number']
        # Now get classes
        cursor.execute("""
            SELECT class_ID, subject, semester FROM classes_database
            WHERE TRIM(teacher_ID) = TRIM(%s)
        """, (prof_id,))
        classes = [
            {
                "class_ID": row['class_ID'],
                "subject": row['subject'],
                "semester": row['semester']
            }
            for row in cursor.fetchall()
        ]
        cursor.close()
        conn.close()
        return jsonify(classes)

    @abstractmethod
    @user_bp.route("/get_subjects")
    def get_subjects():
        class_id = request.args.get("class_id")
        print("Received class_id for /get_subjects:", class_id)
        if not class_id:
            return jsonify([])
        conn = DB.get_mysql_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT DISTINCT subject FROM classes_database WHERE class_ID = %s", (class_id,))
        subjects = [row['subject'] for row in cursor.fetchall()]
        cursor.close()
        conn.close()
        return jsonify(subjects)


    @abstractmethod
    @user_bp.route("/get_sections")
    def get_sections():
      class_id = request.args.get("class_id")
      if not class_id:
          return jsonify([])
      conn = DB.get_mysql_connection()
      cursor = conn.cursor(dictionary=True)
      # Get all student_IDs enrolled in this class
      cursor.execute("SELECT student_ID FROM enrollment_database WHERE class_ID = %s", (class_id,))
      student_ids = [row['student_ID'] for row in cursor.fetchall()]
      sections = []
      if student_ids:
          format_strings = ','.join(['%s'] * len(student_ids))
          cursor.execute(
              f"SELECT DISTINCT course, year_section FROM login_main WHERE id_number IN ({format_strings})",
              tuple(student_ids)
          )
          for row in cursor.fetchall():
              course = row['course'] or ''
              year_section = row['year_section'] or ''
              section = f"{course} {year_section}".strip()
              if section and section not in sections:
                  sections.append(section)
      cursor.close()
      conn.close()
      return jsonify(sections)


    @abstractmethod
    @user_bp.route("/get_students")
    def get_students():
      class_id = request.args.get("class_id")
      section = request.args.get("section")
      if not class_id or not section:
          return jsonify({})  # Return empty dict for grouping

      # Parse section into course and year_section
      section = section.strip()
      if " " in section:
          course, year_section = section.split(" ", 1)
          course = course.strip()
          year_section = year_section.strip()
      else:
          course = section
          year_section = ""

      conn = DB.get_mysql_connection()
      cursor = conn.cursor(dictionary=True)
      # Get all student_IDs enrolled in this class
      cursor.execute("SELECT student_ID FROM enrollment_database WHERE class_ID = %s", (class_id,))
      student_ids = [row['student_ID'] for row in cursor.fetchall()]
      if not student_ids:
          cursor.close()
          conn.close()
          return jsonify({})

      format_strings = ','.join(['%s'] * len(student_ids))
      cursor.execute(
          f"""SELECT name, id_number, course, year_section, profile_pic 
              FROM login_main 
              WHERE id_number IN ({format_strings}) 
              AND course = %s AND year_section = %s""",
          tuple(student_ids) + (course, year_section)
      )
      students = cursor.fetchall()
      cursor.close()
      conn.close()

      # Group students by first letter of last name
      grouped = {}
      for stu in students:
          last_name = stu['name'].split()[-1].upper()
          initial = last_name[0]
          if initial not in grouped:
              grouped[initial] = []
          grouped[initial].append(stu)
      for initial in grouped:
          grouped[initial].sort(key=lambda s: (s['name'].split()[-1].upper(), s['name'].upper()))
      grouped_sorted = dict(sorted(grouped.items()))
      return jsonify(grouped_sorted)


    # @abstractmethod
    # @user_bp.route("/issue_ticket", methods=["POST"])
    # def issue_ticket():
    #     if 'user' not in session:
    #         return redirect(url_for('login'))
    #     subject = request.form.get("subject")
    #     description = request.form.get("description")
    #     user_name = session['user']
    #     # Get user's email
    #     conn = DB.get_mysql_connection()
    #     cursor = conn.cursor(dictionary=True)
    #     cursor.execute("SELECT email FROM login_main WHERE name = %s", (user_name,))
    #     user = cursor.fetchone()
    #     user_email = user['email'] if user else None
    #     # Store ticket in DB (create a table 'issue_tickets' if not exists)
    #     cursor.execute("""
    #         INSERT INTO issue_tickets (user_name, user_email, subject, description, status)
    #         VALUES (%s, %s, %s, %s, %s)
    #     """, (user_name, user_email, subject, description, "Open"))
    #     conn.commit()
    #     cursor.close()
    #     conn.close()
    #     # Redirect with ticket_success=1 so the modal shows
    #     return redirect(url_for('User.profUI', ticket_success=1))


    @abstractmethod
    @user_bp.route('/upload_profile_pic', methods=['POST'])
    def upload_profile_pic():
        if 'user' not in session:
            print("No user in session")
            return redirect(url_for('Login.login'))
        file = request.files['profile_pic']
        print("File received:", file)
        if file:
            filename = secure_filename(f"{session['user']}_{int(time())}_{file.filename}")
            # Use absolute path for the static/profile_pics directory
            static_folder = os.path.join(Main.app.root_path, 'static', 'profile_pics')
            os.makedirs(static_folder, exist_ok=True)
            filepath = os.path.join(static_folder, filename)
            file.save(filepath)
            print("Saved file to:", filepath)
            conn = DB.get_mysql_connection()
            cursor = conn.cursor()
            id_number = session['id_number']
            print("Updating DB for id_number:", id_number, "with filename:", filename)
            cursor.execute("UPDATE login_main SET profile_pic = %s WHERE id_number = %s", (filename, id_number))
            conn.commit()
            cursor.close()
            conn.close()
        else:
            print("No file uploaded!")
        return redirect(url_for('User.profUI'))


#########################################################################################################################
#   LOGIN & REGISTRATION

class Login():
    login_bp = Blueprint("Login", __name__)

    @login_bp.route("/", methods=["GET", "POST"])
    def login():
        """
        Handles user login.
        - On GET: renders the login form.
        - On POST: authenticates user and redirects based on role.
        """
        error = None
        remembered_id = request.cookies.get('remember_id', '')
        remembered_password = request.cookies.get('remember_password', '')
        if request.method == "POST":
            id_number = request.form.get("identification", "").strip()
            password = request.form.get("password", "").strip()
            remember = request.form.get("remember_me")

            conn = DB.get_mysql_connection()
            cursor = conn.cursor(dictionary=True)
            cursor.execute("SELECT * FROM login_main WHERE id_number = %s", (id_number,))
            user = cursor.fetchone()
            cursor.close()
            conn.close()

            if user and user['password'] == password:
                session['user'] = user['name']
                session['id_number'] = user['id_number']
                resp = make_response(
                    redirect(
                        url_for('User.adminUI') if "AD" in id_number else
                        url_for('User.profUI') if "TC" in id_number else
                        url_for('User.studentUI') if "SD" in id_number else
                        url_for('Login.login')
                    )
                )
                # Handle "Remember Me" cookies
                if remember:
                    resp.set_cookie('remember_id', id_number, max_age=60*60*24*30)
                    resp.set_cookie('remember_password', password, max_age=60*60*24*30)
                else:
                    resp.set_cookie('remember_id', '', expires=0)
                    resp.set_cookie('remember_password', '', expires=0)
                return resp
            elif user:
                error = "Incorrect password."
            else:
                error = "Account not found."
        return render_template("login.html", error=error, remembered_id=remembered_id, remembered_password=remembered_password)


    @login_bp.route("/forgot-password", methods=["GET", "POST"])
    def forgot_password():
        error = None
        show_otp = False
        id_number = ""
        if request.method == "POST":
            id_number = request.form.get("identification", "").strip()
            conn = DB.get_mysql_connection()
            cursor = conn.cursor(dictionary=True)
            cursor.execute("SELECT 1 FROM login_main WHERE id_number = %s", (id_number,))
            found = cursor.fetchone() is not None
            cursor.close()
            conn.close()
            if found:
                show_otp = True
            else:
                error = "No such ID number is registered."
        return render_template("forgot_password.html", show_otp=show_otp, error=error, id_number=id_number)


    @login_bp.route('/check_id', methods=['POST'])
    def check_id():
        id_number = request.json.get('id_number')
        conn = DB.get_mysql_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT email FROM login_main WHERE id_number = %s", (id_number,))
        row = cursor.fetchone()
        cursor.close()
        conn.close()
        if row:
            email = row['email']
            otp_code = OTP.generate_otp()
            session['otp'] = otp_code
            session['email'] = email
            session['id_number'] = id_number   
            OTP.send_otp_email(email, otp_code)
            return jsonify({'exists': True, 'email': email})
        else:
            return jsonify({'exists': False})

    @login_bp.route('/change-password', methods=['GET', 'POST'])
    def change_password():
        error = None
        if request.method == 'POST':
            new_password = request.form.get('new-password')
            confirm_password = request.form.get('confirm-password')
            if new_password != confirm_password:
                error = "Passwords do not match."
            else:
                id_number = session.get('id_number') 
                try:
                    conn = DB.get_mysql_connection()
                    cursor = conn.cursor()
                    cursor.execute("UPDATE login_main SET password = %s WHERE id_number = %s", (new_password, id_number))
                    conn.commit()
                    cursor.close()
                    conn.close()
                    session.pop('otp', None)
                    session.pop('email', None)
                    session.pop('id_number', None)
                    return redirect(url_for('Login.changepasswordcomplete'))
                except mysql.connector.Error:
                    error = "Cannot connect to the database. Please try again later."
        return render_template('change_password.html', error=error)

    @login_bp.route('/check_otp', methods=['POST'])
    def check_otp():
        input_otp = request.json.get('otp')
        if input_otp == session.get('otp'):
            return jsonify({'success': True})
        else:
            return jsonify({'success': False, 'error': 'Incorrect OTP. Please try again.'})

    @login_bp.route('/changepasswordcomplete')
    def changepasswordcomplete():
        return render_template('changepasswordcomplete.html')

    @login_bp.route("/credentials", methods=["GET", "POST"]) 
    def register():
        """
        Handles the role selection step of registration.
        - On GET: renders the role selection form.
        - On POST: retrieves the selected role and passes it to the next step.
        """
        
        role = None
        error = None
        if request.method == "POST":
            # Get the selected role from the form submission
            role = request.form.get("role")
        # Render the registration form, passing the selected role and any error
        return render_template("register.html", role=role, error=error)

    @login_bp.route("/register2", methods=["GET", "POST"])
    def register2():
        error = None
        role = request.form.get("role") or session.get("role")
        if request.method == "POST":
            first_name = request.form.get("first_name", "").strip()
            middle_name = request.form.get("middle_name", "").strip()
            last_name = request.form.get("last_name", "").strip()
            id_number = request.form.get("id-number", "").strip()
            input_name = f"{first_name} {middle_name} {last_name}".replace("  ", " ").strip().lower()

            if role and role.lower() == "teacher":
                # Only check ID format for professors
                if "TC" not in id_number:
                    error = "Professor ID must contain 'TC'."
                elif len(id_number) != 15:
                    error = "Invalid ID number."
                else:
                    # Check for duplicate TC ID or name in login_main
                    conn = DB.get_mysql_connection()
                    cursor = conn.cursor(dictionary=True)
                    cursor.execute("SELECT 1 FROM login_main WHERE id_number = %s", (id_number,))
                    if cursor.fetchone():
                        error = "This professor ID is already registered."
                    else:
                        cursor.execute("SELECT 1 FROM login_main WHERE LOWER(name) = %s", (input_name,))
                        if cursor.fetchone():
                            error = "A professor with this name is already registered."
                    cursor.close()
                    conn.close()
            else:
                # ...existing student logic...
                conn = DB.get_mysql_connection()
                cursor = conn.cursor()
                cursor.execute("SELECT 1 FROM login_main WHERE ID_number = %s", (id_number,))
                if cursor.fetchone():
                    error = "This student ID is already registered."
                else:
                    cursor.execute("SELECT name FROM enrollment_database WHERE student_ID = %s", (id_number,))
                    row = cursor.fetchone()
                    if not row:
                        error = "This student ID is not enrolled in the school."
                    else:
                        db_name = row[0].strip().lower()
                        if db_name != input_name:
                            error = "The name you entered does not match our enrollment records."
                cursor.close()
                conn.close()

            if error:
                return render_template("register.html", role=role, error=error)

            session['name'] = f"{first_name} {middle_name} {last_name}".replace("  ", " ").strip()
            session['id_number'] = id_number
            session['role'] = role
            return redirect(url_for('register2_details'))

        return render_template("register2.html", role=role)


    @login_bp.route("/register_details", methods=["GET", "POST"])
    def register2_details():
        """
        Handles the second registration step:
        - For students: requires all fields.
        - For teachers: only requires email and password, and enforces gmail.com domain.
        """
        
        error = None
        if request.method == "POST":
            # Get form data
            email = request.form.get("email", "").strip()
            course = request.form.get("course", "").strip()
            year_section = request.form.get("year_section", "").strip()
            password = request.form.get("password", "").strip()
            confirm_password = request.form.get("confirm_password", "").strip()
            role = session.get('role')

            # Teacher/professor logic
            if role and str(role).lower() in ["teacher", "professor"]:
                # Enforce gmail.com domain for teachers/professors
                if not email.lower().endswith("@gmail.com"):
                    error = "Professors must use a gmail.com email address."
                # Set course and year_section to None for teachers/professors
                course = None
                year_section = None
                # Only require email, password, confirm_password for teachers
                if not all([email, password, confirm_password]):
                    error = "All fields are required."
            else:
                # For students, require all fields
                if not all([email, course, year_section, password, confirm_password]):
                    error = "All fields are required."

            # Check if passwords match
            if not error and password != confirm_password:
                error = "Passwords do not match."
                
            # If any error, re-render the form with error message
            if error:
                return render_template("register2.html", role=session.get('role'), error=error)

            # Store registration details in session for next step
            session['email'] = email
            session['course'] = course
            session['year_section'] = year_section
            session['password'] = password
            # Redirect to OTP verification
            return redirect(url_for('verify'))


        # Render the registration details form
        return render_template("register2.html", role=session.get('role'))

    @login_bp.route('/otp', methods=['GET', 'POST'])
    def verify():
        """
        Handles OTP verification.
        - On GET: generates and sends OTP to the user's email.
        - On POST: checks the submitted OTP against the session value.
        - After 3 wrong attempts, user must wait (cooldown) before retrying.
        - Cooldown increases by 1 minute every 3 failed attempts.
        """
        # Initialize attempt and cooldown tracking in session if not present
        if 'otp_attempts' not in session:
            session['otp_attempts'] = 0
        if 'otp_cooldown_until' not in session:
            session['otp_cooldown_until'] = None
        if 'otp_cooldown_level' not in session:
            session['otp_cooldown_level'] = 1  # in minutes

        now = datetime.now()
        cooldown_until = session.get('otp_cooldown_until')
        cooldown_active = False

        # Check if cooldown is active
        if cooldown_until:
            cooldown_until_dt = datetime.fromisoformat(cooldown_until)
            if now < cooldown_until_dt:
                cooldown_active = True
                remaining = (cooldown_until_dt - now).seconds
            else:
                # Cooldown expired, reset attempts and cooldown
                session['otp_attempts'] = 0
                session['otp_cooldown_until'] = None

        if request.method == 'POST':
            if cooldown_active:
                # Still in cooldown, don't allow attempts
                mins = (remaining // 60) + 1
                return render_template("verify.html", email=session.get('email'),
                                    error=f"Too many incorrect attempts. Please wait {mins} minute(s) before trying again.")
            input_otp = request.form.get('otp')
            if input_otp:
                if input_otp == session.get('otp'):
                    # Success: reset attempts and cooldown
                    session['otp_attempts'] = 0
                    session['otp_cooldown_level'] = 10
                    session['otp_cooldown_until'] = None
                    return redirect(url_for('complete_registration'))
                else:
                    # Wrong OTP: increment attempts
                    session['otp_attempts'] += 1
                    if session['otp_attempts'] >= 3:
                        # Set cooldown: increases by 1 min each time
                        cooldown_minutes = session.get('otp_cooldown_level', 10)
                        session['otp_cooldown_until'] = (now + timedelta(minutes=cooldown_minutes)).isoformat()
                        session['otp_cooldown_level'] = cooldown_minutes + 10
                        return render_template("verify.html", email=session.get('email'),
                                            error=f"Too many incorrect attempts. Please wait {cooldown_minutes} minute(s) before trying again.")
                    else:
                        return render_template("verify.html", email=session.get('email'),
                                            error="Incorrect OTP. Please try again.")
        else:
            # GET request: generate and send a new OTP to the user's email if not in cooldown
            if cooldown_active:
                mins = (remaining // 60) + 1
                return render_template("verify.html", email=session.get('email'),
                                    error=f"Too many incorrect attempts. Please wait {mins} minute(s) before trying again.")
            email = session.get('email')
            if email:
                otp_code = OTP.generate_otp()
                session['otp'] = otp_code
                OTP.send_otp_email(email, otp_code)
                print("OTP sent to", email, "with code", otp_code)

        # Render the OTP verification page (with or without error)
        return render_template("verify.html", email=session.get('email'))


    @login_bp.route('/complete_registration', methods=['GET', 'POST'])
    def complete_registration():
        """
        Finalizes registration:
        - Students and Professors: inserts into login_main.
        """
        # Retrieve registration data from session
        name = session.get('name')
        email = session.get('email')
        role = session.get('role')
        password = session.get('password')
        id_number = session.get('id_number')
        course = session.get('course')
        year_section = session.get('year_section')

        # Normalize role to match ENUM in DB
        role_map = {"student": "Student", "teacher": "Teacher", "admin": "Admin"}
        role = role_map.get(str(role).lower(), role)

        conn = DB.get_mysql_connection()
        cursor = conn.cursor()
        error = None

        try:
            cursor.execute("""
                INSERT INTO login_main (name, id_number, role, password, email, course, year_section)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, (name, id_number, role, password, email, course, year_section))
            conn.commit()
        except mysql.connector.Error as err:
            print("MySQL Error:", err)
            error = f"MySQL Error: {err}"

        cursor.close()
        conn.close()
        return render_template("registration_complete.html", role=role, error=error)

    @login_bp.route("/step2")
    def step2():
        return render_template("role.html")

#########################################################################################################################


#########################################################################################################################
#   FOR ADMIN 

class Admin(User):

    @User.user_bp.route("/admin")
    def adminUI():
        return Admin.get_UI()

    def get_UI():
        conn = DB.get_mysql_connection()
        cursor = conn.cursor(dictionary=True)
        # Summary counts
        cursor.execute("SELECT COUNT(*) FROM login_main WHERE role='Teacher'")
        prof_count = cursor.fetchone()['COUNT(*)']
        cursor.execute("SELECT COUNT(*) FROM login_main WHERE role='Student'")
        student_count = cursor.fetchone()['COUNT(*)']
        cursor.execute("SELECT COUNT(*) FROM classes_database")
        class_count = cursor.fetchone()['COUNT(*)']

        # Classes info for the table (now includes subject)
        cursor.execute("""
            SELECT c.class_ID, c.subject, c.semester, c.teacher_ID, c.status,
                t.name AS professor_name
            FROM classes_database c
            LEFT JOIN login_main t ON c.teacher_ID = t.id_number
        """)
        classes = cursor.fetchall()

        # For each class, get students (with placeholder avatar)
        for cls in classes:
            cursor.execute("""
                SELECT l.name FROM enrollment_database e
                JOIN login_main l ON e.student_ID = l.id_number
                WHERE e.class_ID = %s
            """, (cls['class_ID'],))
            students = cursor.fetchall()
            for stu in students:
                stu['avatar_url'] = url_for('static', filename='images/profile.png')
            cls['students'] = students
            cls['student_count'] = len(students)

        cursor.execute("SELECT id_number, name, role, email FROM login_main")
        all_users = cursor.fetchall()
    
        cursor.execute("SELECT DISTINCT class_ID, subject FROM classes_database")
        class_subjects = cursor.fetchall()
        class_id_subject_map = {row['class_ID']: row['subject'] for row in class_subjects}

        cursor.execute("SELECT * FROM issue_tickets ORDER BY created_at DESC")
        tickets = cursor.fetchall()
        
        cursor.close()
        conn.close()

        
        return render_template(
            "adminUI.html",
            prof_count=prof_count,
            student_count=student_count,
            class_count=class_count,
            classes=classes,
            all_users=all_users,
            class_id_subject_map=class_id_subject_map,
            year=datetime.now().year,
            tickets=tickets
        )

    @User.user_bp.route("/enroll_student", methods=["POST"])
    def enroll_student():
        first_name = request.form.get("first_name", "").strip()
        middle_name = request.form.get("middle_name", "").strip()
        last_name = request.form.get("last_name", "").strip()
        class_ids = request.form.getlist("class_id")  # getlist for multiple
        student_id = request.form.get("student_id", "").strip()
        full_name = f"{first_name} {middle_name} {last_name}".replace("  ", " ").strip()

        conn = DB.get_mysql_connection()
        cursor = conn.cursor()

        # Get the last enrollment_ID and increment it
        cursor.execute("SELECT MAX(enrollment_ID) FROM enrollment_database")
        last_id = cursor.fetchone()[0]
        if last_id and last_id.startswith("ENR-"):
            last_num = int(last_id.split('-')[1])
        else:
            last_num = 0

        for class_id in class_ids:
            new_id = f"ENR-{last_num+1:04d}"
            cursor.execute(
                "INSERT INTO enrollment_database (enrollment_ID, name, student_ID, class_ID) VALUES (%s, %s, %s, %s)",
                (new_id, full_name, student_id, class_id)
            )
            last_num += 1

        conn.commit()
        cursor.close()
        conn.close()
        return redirect(url_for('User.adminUI', enroll='success'))

    @User.user_bp.route("/delete_registrant", methods=["POST"])
    def delete_registrant():
        user_id = request.form.get("user_id")
        if not user_id:
            return redirect(url_for('User.adminUI'))
        conn = DB.get_mysql_connection()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM login_main WHERE id_number = %s", (user_id,))
        conn.commit()
        cursor.close()
        conn.close()
        return redirect(url_for('User.adminUI', remove='success')) 

    @User.user_bp.route("/add_class", methods=["POST"])
    def add_class():
        class_id = request.form.get("class_id", "").strip()
        subject = request.form.get("subject", "").strip()
        semester = request.form.get("semester", "").strip()
        teacher_id = request.form.get("teacher_id", "").strip()

        # Check for duplicate class_id + teacher_id
        conn = DB.get_mysql_connection()
        cursor = conn.cursor()
        cursor.execute(
            "SELECT 1 FROM classes_database WHERE class_ID = %s AND teacher_ID = %s",
            (class_id, teacher_id)
        )
        exists = cursor.fetchone()
        if exists:
            cursor.close()
            conn.close()
            return redirect(url_for('User.adminUI', addclass='duplicate'))

        cursor.execute(
            "INSERT INTO classes_database (class_ID, subject, semester, teacher_ID, status) VALUES (%s, %s, %s, %s, %s)",
            (class_id, subject, semester, teacher_id, "Ongoing")
        )
        conn.commit()
        cursor.close()
        conn.close()
        return redirect(url_for('User.adminUI', addclass='success'))

    @User.user_bp.route("/delete_class", methods=["POST"])
    def delete_class():
        class_id = request.form.get("class_id")
        if not class_id:
            return redirect(url_for('User.adminUI'))
        conn = DB.get_mysql_connection()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM classes_database WHERE class_ID = %s", (class_id,))
        conn.commit()
        cursor.close()
        conn.close()
        return redirect(url_for('User.adminUI'))

    @User.user_bp.route('/change_password', methods=['POST'])
    def change_password_logged_in():
        if 'user' not in session:
            return jsonify({'error': 'Not logged in'}), 401

        data = request.get_json()
        current_password = data.get('current_password')
        new_password = data.get('new_password')
        confirm_password = data.get('confirm_password')

        conn = DB.get_mysql_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT password, role FROM login_main WHERE name = %s", (session['user'],))
        user = cursor.fetchone()
        if not user:
            error = "User not found."
            cursor.close()
            conn.close()
            return jsonify({'error': error}), 400
        elif user['password'] != current_password:
            error = "Current password is incorrect."
            cursor.close()
            conn.close()
            return jsonify({'error': error}), 400
        elif new_password != confirm_password:
            error = "New passwords do not match."
            cursor.close()
            conn.close()
            return jsonify({'error': error}), 400
        else:
            cursor.execute("UPDATE login_main SET password = %s WHERE name = %s", (new_password, session['user']))
            conn.commit()
            role = user.get('role', '').lower()
            cursor.close()
            conn.close()
            # Redirect based on role
            if role == "student":
                return jsonify({'success': "Password changed successfully.", 'redirect': url_for('User.studentUI')})
            elif role == "teacher":
                return jsonify({'success': "Password changed successfully.", 'redirect': url_for('User.profUI')})
            elif role == "admin":
                return jsonify({'success': "Password changed successfully.", 'redirect': url_for('User.adminUI')})
            else:
                return jsonify({'success': "Password changed successfully.", 'redirect': url_for('Login.login')})

        data = request.get_json()
        current_password = data.get('current_password')
        new_password = data.get('new_password')
        confirm_password = data.get('confirm_password')

        conn = DB.get_mysql_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT password FROM login_main WHERE name = %s", (session['user'],))
        user = cursor.fetchone()
        if not user:
            error = "User not found."
            cursor.close()
            conn.close()
            return jsonify({'error': error}), 400
        elif user['password'] != current_password:
            error = "Current password is incorrect."
            cursor.close()
            conn.close()
            return jsonify({'error': error}), 400
        elif new_password != confirm_password:
            error = "New passwords do not match."
            cursor.close()
            conn.close()
            return jsonify({'error': error}), 400
        else:
            cursor.execute("UPDATE login_main SET password = %s WHERE name = %s", (new_password, session['user']))
            conn.commit()
            cursor.close()
            conn.close()
            return jsonify({'success': "Password changed successfully."})

    @staticmethod
    def send_ticket_reply_email(recipient, subject, reply):
        msg = Message(
            subject=f"PyTracker Issue Ticket Reply: {subject}",
            sender=("Pytracker", Main.app.config['MAIL_USERNAME']),
            recipients=[recipient]
        )
        msg.body = f"Admin replied to your issue:\n\n{reply}\n\nThank you for using PyTracker."
        Main.mail.send(msg)

    @User.user_bp.route("/reply_ticket", methods=["POST"])
    def reply_ticket():
        ticket_id = request.form.get("ticket_id")
        reply = request.form.get("reply")
        conn = DB.get_mysql_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT user_email, subject FROM issue_tickets WHERE ticket_id = %s", (ticket_id,))
        ticket = cursor.fetchone()
        if ticket:
            Admin.send_ticket_reply_email(ticket['user_email'], ticket['subject'], reply)
            cursor.execute("UPDATE issue_tickets SET admin_reply = %s, status = %s WHERE ticket_id = %s", (reply, "Closed", ticket_id))
            conn.commit()
        cursor.close()
        conn.close()
        return redirect(url_for('User.adminUI', reply_success=1))

    @User.user_bp.route("/issue_tickets")
    def issue_tickets():
        conn = DB.get_mysql_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM issue_tickets ORDER BY created_at DESC")
        tickets = cursor.fetchall()
        cursor.close()
        conn.close()
        return render_template("issue_tickets.html", tickets=tickets)



#########################################################################################################################

#########################################################################################################################
#   PROFESSOR'S UI

class Prof(User):
    @User.user_bp.route("/prof")
    def profUI():
        return Prof.get_UI()

    def get_UI():
        prof_name = None
        prof_role = None
        prof_pic = None
        attendance_calendar = {}
        histogram_counts = {
            "Absent": 0,
            "Late": 0,
            "No Classes/Excused": 0,
            "Attended": 0,
            "Days Remaining": 0
        }
        today = datetime.today()
        school_year_start = datetime(today.year if today.month >= 6 else today.year - 1, 6, 1)
        school_days_passed = 0
        total_school_days = 180

        d = school_year_start
        while d <= today:
            if d.weekday() < 5:
                school_days_passed += 1
            d += timedelta(days=1)
        days_remaining = max(0, total_school_days - school_days_passed)
        histogram_counts["Days Remaining"] = days_remaining

        months = []
        for i in range(-2, 6):
            month = (today.month + i - 1) % 12 + 1
            year = today.year + ((today.month + i - 1) // 12)
            months.append((year, month))

        if 'user' in session:
            conn = DB.get_mysql_connection()
            cursor = conn.cursor(dictionary=True)
            cursor.execute("SELECT name, id_number, role, profile_pic FROM login_main WHERE id_number = %s", (session['id_number'],))
            prof_row = cursor.fetchone()
            prof_name = prof_row['name'] if prof_row else None
            prof_role = prof_row['role'] if prof_row else None
            prof_pic = prof_row['profile_pic'] if prof_row else None
            prof_id_test = prof_row['id_number'] if prof_row else None
            class_ids = None
            if prof_id_test:
                cursor.execute("SELECT class_ID FROM classes_database WHERE teacher_ID = %s", (prof_id_test,))
                class_rows = cursor.fetchall()
                class_ids = [row['class_ID'] for row in class_rows] if class_rows else None
            cursor.close()
            conn.close()
        return render_template(
            "profUI.html",
            prof_name=prof_name,
            prof_role=prof_role,
            prof_pic=prof_pic,
            prof_id=prof_id_test,
            attendance_calendar=attendance_calendar,
            histogram_counts=histogram_counts,
            datetime=datetime,
            calendar=calendar
        )

    @User.user_bp.route('/admin_secret_check', methods=['POST'])
    def admin_secret_check():
        data = request.get_json()
        if data.get('secret') == 'supersecret':
            session['user'] = 'admin'
            return jsonify({'success': True})
        else:
            return jsonify({'success': False})


class Student(User):
    
    @User.user_bp.route("/student")
    def studentUI():
        return Student.get_UI()
    
    def get_UI():
        student_name = None
        student_role = None
        student_pic = None
        student_id = None
        student_classes = []
        student_cys = None
        if 'id_number' in session:
            conn = DB.get_mysql_connection()
            cursor = conn.cursor(dictionary=True)
            cursor.execute("SELECT name, role, profile_pic, id_number, course, year_section FROM login_main WHERE id_number = %s", (session['id_number'],))
            row = cursor.fetchone()
            if row:
                student_name = row['name']
                student_role = row['role']
                student_pic = row['profile_pic']
                student_id = row['id_number']
                student_cys = f"{row['course']} {row['year_section']}".strip()
            # Get all classes the student is enrolled in
            cursor.execute("SELECT class_ID FROM enrollment_database WHERE student_ID = %s", (session['id_number'],))
            student_classes = [r['class_ID'] for r in cursor.fetchall()]
            cursor.close()
            conn.close()
        return render_template(
            "studentUI.html",
            student_name=student_name,
            student_role=student_role,
            prof_pic=student_pic,
            student_id=student_id,
            student_classes=student_classes,
            student_cys=student_cys
        )

class Funcs:
    funcs_bp = Blueprint("Funcs", __name__)

    @staticmethod
    @funcs_bp.route("/prof_dropdown_data")
    def prof_dropdown_data():
        if 'user' not in session:
            return jsonify({})
        conn = DB.get_mysql_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT id_number FROM login_main WHERE name = %s", (session['user'],))
        prof = cursor.fetchone()
        if not prof:
            cursor.close()
            conn.close()
            return jsonify({})
        prof_id = prof['id_number']
        cursor.execute("SELECT class_ID, subject, semester FROM classes_database WHERE teacher_ID = %s", (prof_id,))
        classes = cursor.fetchall()
        class_sections = {}
        for cls in classes:
            cursor.execute(
                "SELECT DISTINCT course, year_section FROM enrollment_database JOIN login_main ON enrollment_database.student_ID = login_main.id_number WHERE class_ID = %s",
                (cls['class_ID'],)
            )
            sections = [f"{row['course']} {row['year_section']}".strip() for row in cursor.fetchall()]
            class_sections[cls['class_ID']] = sections
        cursor.close()
        conn.close()
        return jsonify({
            "classes": classes,
            "class_sections": class_sections
        })

    from datetime import datetime


    @staticmethod
    @funcs_bp.route("/mark_attendance_bulk", methods=["POST"])
    def mark_attendance_bulk():
        data = request.get_json()
        records = data.get("records", [])
        if not records:
            return jsonify({"success": False, "error": "No records provided"})
        conn = DB.get_mysql_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT MAX(`control_no`) FROM attendance_database")
        max_control = cursor.fetchone()[0] or 0
        control_no = max_control + 1

        class_id = records[0]['class_ID']
        cursor.execute("SELECT subject FROM classes_database WHERE class_ID = %s", (class_id,))
        subject = cursor.fetchone()[0]
        section = records[0].get('section', None)
        prof_id = session.get('id_number')
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        manila = pytz.timezone('Asia/Manila')
        for rec in records:
            now_str = datetime.now(manila).strftime("%Y-%m-%d %H:%M:%S")
            cursor.execute(
                "INSERT INTO attendance_database (`control_no`, date, class_ID, student_ID, status) VALUES (%s, %s, %s, %s, %s)",
                (control_no, now_str, rec['class_ID'], rec['student_ID'], rec['status'])
            )
            # Use now_str for the notification message as well:
            student_msg = f"Your attendance was recorded for {subject} ({section}) on {now_str} as {rec['status']}."
            cursor.execute(
                "INSERT INTO notifications (user_id, message, created_at) VALUES (%s, %s, %s)",
                (rec['student_ID'], student_msg, now_str)
            )
            control_no += 1

        prof_msg = f"You took attendance for {subject} ({section}) on {now}."
        cursor.execute(
            "INSERT INTO notifications (user_id, message, created_at) VALUES (%s, %s, %s)",
            (prof_id, prof_msg, now)
        )

        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({"success": True})

    @staticmethod
    @funcs_bp.route("/get_notifications")
    def get_notifications():
        if 'id_number' not in session:
            return jsonify([])
        user_id = session['id_number']
        conn = DB.get_mysql_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM notifications WHERE user_id = %s ORDER BY created_at DESC LIMIT 20", (user_id,))
        notifications = cursor.fetchall()
        cursor.close()
        conn.close()
        return jsonify(notifications)

    @staticmethod
    @funcs_bp.route("/issue_ticket", methods=["POST"])
    def issue_ticket():
        if 'user' not in session:
            return redirect(url_for('Login.login'))
        subject = request.form.get("subject")
        description = request.form.get("description")
        user_name = session['user']
        conn = DB.get_mysql_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT email, role, id_number FROM login_main WHERE name = %s", (user_name,))
        user = cursor.fetchone()
        user_email = user['email'] if user else None
        user_role = user['role'].lower() if user and 'role' in user else None
        user_id = user['id_number'] if user and 'id_number' in user else None
        cursor.execute("""
            INSERT INTO issue_tickets (user_name, user_email, subject, description, status)
            VALUES (%s, %s, %s, %s, %s)
        """, (user_name, user_email, subject, description, "Open"))
        conn.commit()
        cursor.close()
        conn.close()
        # Redirect based on role
        if user_role == "student":
            return redirect(url_for('User.studentUI', ticket_success=1))
        elif user_role == "teacher":
            return redirect(url_for('User.profUI', ticket_success=1))
        elif user_role == "admin":
            return redirect(url_for('User.adminUI', ticket_success=1))
        else:
            return redirect(url_for('Login.login'))

    @staticmethod
    @funcs_bp.route('/upload_profile_pic', methods=['POST'])
    def upload_profile_pic():
        if 'user' not in session:
            print("No user in session")
            return redirect(url_for('Login.login'))
        file = request.files['profile_pic']
        print("File received:", file)
        if file:
            filename = secure_filename(f"{session['user']}_{int(time.time())}_{file.filename}")
            static_folder = os.path.join(Main.app.root_path, 'static', 'profile_pics')
            os.makedirs(static_folder, exist_ok=True)
            filepath = os.path.join(static_folder, filename)
            file.save(filepath)
            print("Saved file to:", filepath)
            conn = DB.get_mysql_connection()
            cursor = conn.cursor(dictionary=True)
            id_number = session['id_number']
            print("Updating DB for id_number:", id_number, "with filename:", filename)
            cursor.execute("UPDATE login_main SET profile_pic = %s WHERE id_number = %s", (filename, id_number))
            cursor.execute("SELECT role FROM login_main WHERE id_number = %s", (id_number,))
            user = cursor.fetchone()
            role = user['role'].lower() if user and 'role' in user else None
            conn.commit()
            cursor.close()
            conn.close()
        else:
            print("No file uploaded!")
            return redirect(url_for('Login.login'))

        if role == "teacher":
            return redirect(url_for('User.profUI'))
        elif role == "student":
            return redirect(url_for('User.studentUI'))
        elif role == "admin":
            return redirect(url_for('User.adminUI'))
        else:
            return redirect(url_for('Login.login'))

    @staticmethod
    @funcs_bp.route("/get_attendance_summary")
    def get_attendance_summary():
        class_id = request.args.get("class_id")
        section = request.args.get("section")
        if not class_id or not section:
            return jsonify({})
        section = section.strip()
        if " " in section:
            course, year_section = section.split(" ", 1)
            course = course.strip()
            year_section = year_section.strip()
        else:
            course = section
            year_section = ""
        conn = DB.get_mysql_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT student_ID FROM enrollment_database WHERE class_ID = %s", (class_id,))
        student_ids = [row['student_ID'] for row in cursor.fetchall()]
        if not student_ids:
            cursor.close()
            conn.close()
            return jsonify({})
        format_strings = ','.join(['%s'] * len(student_ids))
        cursor.execute(
            f"""SELECT id_number, name FROM login_main 
                WHERE id_number IN ({format_strings}) 
                AND course = %s AND year_section = %s""",
            tuple(student_ids) + (course, year_section)
        )
        students = {row['id_number']: row['name'] for row in cursor.fetchall()}
        cursor.execute(
            f"""SELECT student_ID, status, date FROM attendance_database 
                WHERE class_ID = %s AND student_ID IN ({format_strings})""",
            (class_id, *student_ids)
        )
        summary = {}
        for stu_id in students:
            summary[stu_id] = {"present":0, "absent":0, "late":0, "dates":[]}
        for row in cursor.fetchall():
            stu_id = row['student_ID']
            status = row['status'].strip().lower()
            date = row['date']
            if stu_id in summary:
                if status == "present" or status == "attended":
                    summary[stu_id]["present"] += 1
                elif status == "absent":
                    summary[stu_id]["absent"] += 1
                elif status == "late":
                    summary[stu_id]["late"] += 1
                summary[stu_id]["dates"].append({"status": status.title(), "date": date})
        cursor.close()
        conn.close()
        return jsonify(summary)

    @staticmethod
    @funcs_bp.route("/download_attendance_pdf")
    def download_attendance_pdf():
        class_id = request.args.get("class_id")
        section = request.args.get("section")
        if not class_id or not section:
            return "Missing data", 400

        # Parse section into course and year_section
        section = section.strip()
        if " " in section:
            course, year_section = section.split(" ", 1)
            course = course.strip()
            year_section = year_section.strip()
        else:
            course = section
            year_section = ""

        conn = DB.get_mysql_connection()
        cursor = conn.cursor(dictionary=True)

        # Get all student_IDs enrolled in this class and section
        cursor.execute("SELECT student_ID FROM enrollment_database WHERE class_ID = %s", (class_id,))
        student_ids = [row['student_ID'] for row in cursor.fetchall()]
        if not student_ids:
            cursor.close()
            conn.close()
            return "No students", 404

        format_strings = ','.join(['%s'] * len(student_ids))
        cursor.execute(
            f"""SELECT id_number, name FROM login_main 
                WHERE id_number IN ({format_strings}) 
                AND course = %s AND year_section = %s""",
            tuple(student_ids) + (course, year_section)
        )
        students = {row['id_number']: row['name'] for row in cursor.fetchall()}

        # Get attendance records for these students in this class
        cursor.execute(
            f"""SELECT student_ID, status, date FROM attendance_database 
                WHERE class_ID = %s AND student_ID IN ({format_strings})""",
            (class_id, *student_ids)
        )
        summary = {}
        for stu_id in students:
            summary[stu_id] = {"present":0, "absent":0, "late":0, "dates":[]}
        for row in cursor.fetchall():
            stu_id = row['student_ID']
            status = row['status'].strip().lower()
            date = row['date']
            if stu_id in summary:
                if status == "present" or status == "attended":
                    summary[stu_id]["present"] += 1
                elif status == "absent":
                    summary[stu_id]["absent"] += 1
                elif status == "late":
                    summary[stu_id]["late"] += 1
                summary[stu_id]["dates"].append({"status": status.title(), "date": date})
        cursor.close()
        conn.close()

        # Generate HTML for the table (with colors)
        html = """
        <html>
        <head>
        <meta charset="utf-8">
        <style>
        body {{ font-family: 'Roboto', Arial, sans-serif; }}
        table {{ width: 100%; border-collapse: collapse; min-width: 700px; }}
        th, td {{ padding: 8px; border-bottom: 1px solid #e3e9ed; }}
        th {{ background: #f5f7fa; }}
        .present {{ color: #219653; font-weight: 600; background: #eafaf1; }}
        .absent  {{ color: #d32f2f; font-weight: 600; background: #fdeaea; }}
        .late    {{ color: #b8860b; font-weight: 600; background: #fffbe6; }}
        .dates-cell {{ font-size: 0.95em; }}
        </style>
        </head>
        <body>
        <h2>Attendance Report for Class: {class_id} | Section: {section}</h2>
        <table>
            <thead>
                <tr>
                    <th>Name</th>
                    <th>Present</th>
                    <th>Absent</th>
                    <th>Late</th>
                    <th>Dates</th>
                </tr>
            </thead>
            <tbody>
        """.format(class_id=class_id, section=section)

        for stu_id, name in students.items():
            att = summary[stu_id]
            html += f"""
            <tr>
                <td>{name}</td>
                <td class="present">{att['present']}</td>
                <td class="absent">{att['absent']}</td>
                <td class="late">{att['late']}</td>
                <td class="dates-cell">
                    {''.join([
                        f"<span class='{d['status'].lower()}'>{d['status']}</span>: {d['date']}<br>"
                        for d in att['dates']
                    ])}
                </td>
            </tr>
            """

        html += """
            </tbody>
        </table>
        </body>
        </html>
        """
        
        import io

        pdf_bytes = pdfkit.from_string(html, False, configuration=PDFKIT_CONFIG)
        pdf_io = io.BytesIO(pdf_bytes)
        pdf_io.seek(0)
        filename = f"attendance_{class_id}_{section.replace(' ','_')}.pdf"
        return send_file(pdf_io, mimetype='application/pdf', as_attachment=True, download_name=filename)

    @staticmethod
    @funcs_bp.route("/download_attendance_excel")
    def download_attendance_excel():
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        import io

        class_id = request.args.get("class_id")
        section = request.args.get("section")
        if not class_id or not section:
            return "Missing data", 400

        # Parse section into course and year_section
        section = section.strip()
        if " " in section:
            course, year_section = section.split(" ", 1)
            course = course.strip()
            year_section = year_section.strip()
        else:
            course = section
            year_section = ""

        conn = DB.get_mysql_connection()
        cursor = conn.cursor(dictionary=True)

        # Get all student_IDs enrolled in this class and section
        cursor.execute("SELECT student_ID FROM enrollment_database WHERE class_ID = %s", (class_id,))
        student_ids = [row['student_ID'] for row in cursor.fetchall()]
        if not student_ids:
            cursor.close()
            conn.close()
            return "No students", 404

        format_strings = ','.join(['%s'] * len(student_ids))
        cursor.execute(
            f"""SELECT id_number, name FROM login_main 
                WHERE id_number IN ({format_strings}) 
                AND course = %s AND year_section = %s""",
            tuple(student_ids) + (course, year_section)
        )
        students = {row['id_number']: row['name'] for row in cursor.fetchall()}

        # Get attendance records for these students in this class
        cursor.execute(
            f"""SELECT student_ID, status, date FROM attendance_database 
                WHERE class_ID = %s AND student_ID IN ({format_strings})""",
            (class_id, *student_ids)
        )
        summary = {}
        for stu_id in students:
            summary[stu_id] = {"present":0, "absent":0, "late":0, "dates":[]}
        for row in cursor.fetchall():
            stu_id = row['student_ID']
            status = row['status'].strip().lower()
            date = row['date']
            if stu_id in summary:
                if status == "present" or status == "attended":
                    summary[stu_id]["present"] += 1
                elif status == "absent":
                    summary[stu_id]["absent"] += 1
                elif status == "late":
                    summary[stu_id]["late"] += 1
                summary[stu_id]["dates"].append({"status": status.title(), "date": date})
        cursor.close()
        conn.close()

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance"

        # Header
        ws.append(["Name", "Present", "Absent", "Late", "Dates"])
        header_font = Font(bold=True)
        ws.row_dimensions[1].height = 22
        for col in range(1, 6):
            ws.cell(row=1, column=col).font = header_font
            ws.cell(row=1, column=col).alignment = Alignment(horizontal="center", vertical="center")

        # Color fills
        present_fill = PatternFill(start_color="EAF9F1", end_color="EAF9F1", fill_type="solid")
        absent_fill = PatternFill(start_color="FDEAEA", end_color="FDEAEA", fill_type="solid")
        late_fill = PatternFill(start_color="FFFBE6", end_color="FFFBE6", fill_type="solid")

        # Data rows
        for stu_id, name in students.items():
            att = summary[stu_id]
            dates_str = "\n".join([f"{d['status']}: {d['date']}" for d in att['dates']])
            row = [name, att['present'], att['absent'], att['late'], dates_str]
            ws.append(row)
            r = ws.max_row
            ws.cell(row=r, column=2).fill = present_fill
            ws.cell(row=r, column=3).fill = absent_fill
            ws.cell(row=r, column=4).fill = late_fill
            ws.cell(row=r, column=2).font = Font(color="219653", bold=True)
            ws.cell(row=r, column=3).font = Font(color="D32F2F", bold=True)
            ws.cell(row=r, column=4).font = Font(color="B8860B", bold=True)
            ws.cell(row=r, column=5).alignment = Alignment(wrap_text=True)

        # Adjust column widths
        ws.column_dimensions['A'].width = 28
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 40

        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"attendance_{class_id}_{section.replace(' ','_')}.xlsx"
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        as_attachment=True, download_name=filename)

    @staticmethod
    @funcs_bp.route("/get_notifications_student")
    def get_notifications_student():
        if 'id_number' not in session:
            return jsonify([])
        user_id = session['id_number']
        conn = DB.get_mysql_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM notifications WHERE user_id = %s ORDER BY created_at DESC LIMIT 20", (user_id,))
        notifications = cursor.fetchall()
        cursor.close()
        conn.close()
        return jsonify(notifications)

    @staticmethod
    @funcs_bp.route("/histogram_data_student")
    def histogram_data_student():
        if 'id_number' not in session:
            return jsonify({"error": "Not logged in"}), 401

        student_id = session['id_number']
        conn = DB.get_mysql_connection()
        cursor = conn.cursor(dictionary=True)

        cursor.execute("""
            SELECT status, COUNT(*) as count 
            FROM attendance_database 
            WHERE student_ID = %s 
            GROUP BY status
        """, (student_id,))
        status_counts = {row['status'].strip().lower(): row['count'] for row in cursor.fetchall()}

        absent = status_counts.get('absent', 0)
        late = status_counts.get('late', 0)
        attended = status_counts.get('attended', 0) + status_counts.get('present', 0)

        cursor.execute("SELECT COUNT(DISTINCT date) as days_recorded FROM attendance_database WHERE student_ID = %s", (student_id,))
        days_recorded = cursor.fetchone()['days_recorded']
        days_remaining = max(0, 180 - days_recorded)

        cursor.close()
        conn.close()

        return jsonify({
            "Absent": absent,
            "Late": late,
            "Attended": attended,
            "Days Remaining": days_remaining
        })

    
if __name__ == "__main__":
    main_app = Main()
    main_app.run()